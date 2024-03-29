'''
:title: passdown.py
:version: V1.0
:author: Jack Murray
:company: Edwards Vacuum
:email: jack.murray@edwardsvacuum.com
:requires: openpyxl, datetime, openpyxl.worksheet.table, openpyxl.utils.cell, openpyxl.drawing.image, openpyxl.worksheet.datavalidation, openpyxl.styles, openpyxl.styles.PatternFill, openpyxl.styles.Border, openpyxl.styles.Side, openpyxl.styles.Alignment, openpyxl.styles.Protection, openpyxl.styles.Font
:date: 2023-03-22
'''

#TODO figure out how to get this set up as a windows script
import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
#from openpyxl.utils.cell import range_boundaries
import openpyxl.utils.cell
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# get date range for generating dictionary
start_day = datetime.date(2023, 4, 3)
end_day = datetime.date(2023, 7, 14)
# TODO Get this info from somewhere other than the code

def get_workdays(start_date, end_date):

    '''
    return a list of the dates of all the workdays between start_date and end_date

    :param start_date: a datetime.date object
    :param end_date: a datetime.date object
    :returns: a list of dictionaries in the form {'date': datetime.date, 'week_num': int}
    :rtype: list
    :requires: datetime
    :date: 2023-03-22
    '''

    workdays = []
    current_date = start_date

    while current_date <= end_date:
        if current_date.weekday() < 5:  # Monday = 0, Friday = 4
            workdays.append([current_date,
                                current_date.isocalendar()[1],
                                current_date.strftime("%m-%d")])

        current_date += datetime.timedelta(days=1)

    return workdays

def create_workbook(sheet_names, filename):
    '''
    Create a new excel workbook with the given sheet names

    :param sheet_names: a list of strings containing the names of the sheets to create
    :param filename: a string containing the name of the file to save the workbook to (including the file extension)
    :requires: openpyxl
    :date: 2023-03-22
    '''

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Rename the default sheet to the first sheet name in the list
    workbook.active.title = sheet_names[0]

    # Create the remaining sheets using the remaining sheet names in the list
    for name in sheet_names[1:]:
        workbook.create_sheet(title=name)

    # Save the workbook to a file
    workbook.save(filename)

def get_sheet_names(workdays):
    '''
    Create a list of strings containing the names of the sheets to create

    :param workdays: a list of dictionaries in the form {'date': datetime.date, 'week_num': int}
    :returns: a list of strings containing the names of the sheets to create
    :rtype: list
    :requires: openpyxl
    :date: 2023-03-22
    '''

    sheet_names = []
    sheet_names.append("Contents")
    sheet_names.append("Passdown Summary")
    for day in workdays:
        sheet_names.append(day[2])
    sheet_names.append("passdown_assets")

    return sheet_names

def copy_template_to_sheets(template_filename, sheet_range):
    '''
    Copy the template sheet to the specified sheets in the workbook
    
    :param template_filename: a string containing the name of the file to load the workbook from (including the file extension)
    :param sheet_range: a range of sheet names to copy the template to
    :requires: openpyxl
    :date: 2023-03-22
    '''

    # Load the template file
    wb = openpyxl.load_workbook(template_filename)

    # Get the template sheet
    template_sheet = wb.active

    # Loop over the sheets in the specified range
    for sheet_name in wb.sheetnames[sheet_range]:
        # Skip the template sheet
        if sheet_name == template_sheet.title:
            continue

        # Copy the template sheet to the current sheet
        target_sheet = wb[sheet_name]
        target_sheet.delete_rows(1, target_sheet.max_row)
        for row in template_sheet.iter_rows():
            for cell in row:
                target_sheet[cell.coordinate].value = cell.value
                target_sheet[cell.coordinate].number_format = cell.number_format
                target_sheet[cell.coordinate].font = cell.font
                target_sheet[cell.coordinate].border = cell.border
                target_sheet[cell.coordinate].fill = cell.fill
                target_sheet[cell.coordinate].alignment = cell.alignment

    # Save the modified workbook
    wb.save(template_filename)

def create_daily_sheets(filename, workdays):
    '''
    Create a new excel workbook with the given sheet names

    :param filename: a string containing the name of the file to load the workbook from (including the file extension)
    :param workdays: a list of dictionaries in the form {'date': datetime.date, 'week_num': int}
    :requires: openpyxl
    :date: 2023-03-231
    '''

    # Load the workbook
    wb = openpyxl.load_workbook(filename)

    # Loop over the workdays
    for day in workdays:

        # Get the sheet for the current day
        sheet = wb[day[2]]
        print(day[1])
        #alternating week tab colors
        if day[1] % 2 == 0:
            sheet.sheet_properties.tabColor = "b22222"
        else:
            sheet.sheet_properties.tabColor = "708090"

        #resize the columns
        sheet.column_dimensions["A"].width = 11
        sheet.column_dimensions["B"].width = 17
        sheet.column_dimensions["C"].width = 43
        sheet.column_dimensions["D"].width = 65
        sheet.column_dimensions["E"].width = 8.75
        sheet.column_dimensions["F"].width = 10.5
        sheet.column_dimensions["G"].width = 22.5
        sheet.column_dimensions["H"].width = 22.75
        sheet.column_dimensions["I"].width = 22.75
        sheet.column_dimensions["J"].width = 3.5
        sheet.column_dimensions["K"].width = 11
        sheet.column_dimensions["L"].width = 14.5
        sheet.column_dimensions["M"].width = 23
        sheet.column_dimensions["N"].width = 71
        sheet.column_dimensions["O"].width = 12
        sheet.column_dimensions["P"].width = 34

        #create sheet title
        sheet.merge_cells("A1:P1")
        sheet.row_dimensions[1].height = 84
        sheet["A1"] = "Project Passdown for Work Week " + str(day[1]) + " - " + day[0].strftime("%A, %B %d, %Y")
        sheet["A1"].font = openpyxl.styles.Font(bold=True, size=20)
        sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="bottom")
        sheet["A1"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="C0C0C0")
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=16):
            for cell in row:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))
        #Add the logo
        img = openpyxl.drawing.image.Image(
            '/home/jack/Dropbox/Workspace/code/passdown/python/Edwards_logo_for_sheets.png')
        sheet.add_image(img, "G1")

        # create edwards contact info
        sheet.merge_cells("A2:P2")
        sheet["A2"] = "Edwards Project Manager: Robert Nolan (503)753-0590   -   Edwards Site Manager: Joseph Baca (505)975-9464"
        sheet["A2"].font = openpyxl.styles.Font(size=14)
        sheet["A2"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        sheet["A2"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="C0C0C0")
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=16):
            for cell in row:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))

        #create passdown and look ahead headers
        sheet.merge_cells("A3:I3")
        sheet["A3"] = "Passdown"
        sheet["A3"].font = openpyxl.styles.Font(bold=True, size=16)
        sheet["A3"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        sheet["A3"].border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))
        sheet["A3"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="808080")
        
        sheet.merge_cells("K3:P3")
        sheet["K3"] = "Look Ahead"
        sheet["K3"].font = openpyxl.styles.Font(bold=True, size=16)
        sheet["K3"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        sheet["K3"].border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))
        sheet["K3"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="808080")
        
        #create the divider between passdown and look ahead
        sheet.merge_cells("J3:J103")
        sheet["J3"].fill = openpyxl.styles.PatternFill(
            patternType="solid", fgColor="000000")
        
        #header styles
        

        #create the passdown headers
        sheet["A4"] = "Task ID"
        sheet["B4"] = "Team Assigned"
        sheet["C4"] = "Task"
        sheet["D4"] = "Outcome"
        sheet["E4"] = "SMA"
        sheet["F4"] = "KAWIs"
        sheet["G4"] = "Next Step"
        sheet["H4"] = "AIO Request"
        sheet["I4"] = "AIO Completion"

        # passdown header formatting
        for row in sheet.iter_rows(min_row=4, max_row=4, min_col=1, max_col=9):
            for cell in row:
                cell.font = openpyxl.styles.Font(bold=True, size=11)
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))

        #create the look ahead headers
        sheet["K4"] = "Task ID"
        sheet["L4"] = "Priority Rank"
        sheet["M4"] = "Estimated Completion"
        sheet["N4"] = "Task"
        sheet["O4"] = "Time"

        # Look ahead headers formatting
        for row in sheet.iter_rows(min_row=4, max_row=4, min_col=11, max_col=15):
            for cell in row:
                cell.font = openpyxl.styles.Font(bold=True, size=11)
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))

        #create passdown and look ahead table
        passdown_display_name = day[2] + "_Passdown"
        look_ahead_display_name = day[2] + "_Look_Ahead"
        passdown_table = Table(ref='A4:I103', displayName=passdown_display_name, headerRowCount=1)
        look_ahead_table = Table(ref='K4:O103', displayName=look_ahead_display_name, headerRowCount=1)


        sheet_table_style = TableStyleInfo(
            name="TableStyleMedium15", showRowStripes=True, showColumnStripes=False)
        
        passdown_table.tableStyleInfo = sheet_table_style
        look_ahead_table.tableStyleInfo = sheet_table_style

        sheet.add_table(passdown_table)
        sheet.add_table(look_ahead_table)

        #change header style
        header_style = openpyxl.styles.Font(bold=True, size=11, color="e5e4e2")
        for row in sheet.iter_rows(min_row=4, max_row=4, min_col=1, max_col=15):
            for cell in row:
                cell.font = header_style

        #create passdown column borders
        for col in sheet.iter_cols(min_row=4, max_row=103, min_col=1, max_col=9):
            for cell in col:
                cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style="thick"))

        #create look ahead column borders
        for col in sheet.iter_cols(min_row=4, max_row=103, min_col=11, max_col=16):
            for cell in col:
                cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style="thick"))

        #create bottom boarder
        for row in sheet.iter_rows(min_row=103, max_row=103, min_col=1, max_col=15):
            for cell in row:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick"))
        
        #create right boarder
        for col in sheet.iter_cols(min_row=1, max_row=3, min_col=16, max_col=16):
            for cell in col:
                cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style="thick"))

        # Cell data validation for SMA and KAWIs
        wb_assets = wb["passdown_assets"]
        
        # passdown_assets --> AIO Status
        wb_assets["A1"] = "NOT AIO: Edwards Internal"
        wb_assets["A2"] = "Misc. See Comments"
        wb_assets["A3"] = "Request Per Tool Owner"
        wb_assets["A4"] = "Abatement Commissioning"
        wb_assets["A5"] = "Abatement Demo - Clean"
        wb_assets["A6"] = "Abatement Demo - Like New"
        wb_assets["A7"] = "Abatement Demo - Scrap"
        wb_assets["A8"] = "Abatement POCs"
        wb_assets["A9"] = "Cap Count"
        wb_assets["A10"] = "Chiller Install"
        wb_assets["A11"] = "EPA testing"
        wb_assets["A12"] = "Frame Reassembly"
        wb_assets["A13"] = "Frame Split"
        wb_assets["A14"] = "LSS Termination"
        wb_assets["A15"] = "LSS Testing (SL1)"
        wb_assets["A16"] = "LSS Testing(SL2)"
        wb_assets["A17"] = "Move in and inventory"
        wb_assets["A18"] = "NG turn off"
        wb_assets["A19"] = "NG turn on SL1"
        wb_assets["A20"] = "Pump Commissioning"
        wb_assets["A21"] = "Pump Demo"
        wb_assets["A22"] = "Pump POCs"
        wb_assets["A23"] = "Callback - Abatement POC"
        wb_assets["A24"] = "Callback - Pump POCs"
        wb_assets["A25"] = "Callback - Abatement Commissioning"
        wb_assets["A26"] = "Callback - Pump Commissioning"
        
        # passdown_assets --> AIO completion
        wb_assets["B1"] = "COMPLETE"
        wb_assets["B2"] = "Ongoing"
        wb_assets["B3"] = "Construction Not Complete"
        wb_assets["B4"] = "Damaged Cables"
        wb_assets["B5"] = "Edwards"
        wb_assets["B6"] = "Intermixing"
        wb_assets["B7"] = "Issue With Pressure"
        wb_assets["B8"] = "Kinked Tubing"
        wb_assets["B9"] = "Label Issues"
        wb_assets["B10"] = "Misc. See Comments"
        wb_assets["B11"] = "Punchlist Item Found"
        wb_assets["B12"] = "Abatement Not Commissioned"
        wb_assets["B13"] = "Area RED Taped"
        wb_assets["B14"] = "CAB Failed QAQC"
        wb_assets["B15"] = "CAB Failed Testing"
        wb_assets["B16"] = "Duplicated Request"
        wb_assets["B17"] = "Failed QAQC"
        wb_assets["B18"] = "Failed Terms"
        wb_assets["B19"] = "Incorrect Gas/POC Added"
        wb_assets["B20"] = "Incorrect Information On Form"
        wb_assets["B21"] = "Incorrect Request Form"
        wb_assets["B22"] = "Issue Found At Testing"
        wb_assets["B23"] = "Manufacturing Did Not Release Tool"
        wb_assets["B24"] = "Misc. See Comments"
        wb_assets["B25"] = "No Intel Representative"
        wb_assets["B26"] = "No Longer Needed"
        wb_assets["B27"] = "Per CC's request"
        wb_assets["B28"] = "Per Tool Owner"
        wb_assets["B29"] = "Per Trades Request"
        wb_assets["B30"] = "Requested Before CAB was Completed"
        wb_assets["B31"] = "Test"
        wb_assets["B32"] = "Tool Has NOT met SL1"
        wb_assets["B33"] = "Tool not Construction Complete"
        wb_assets["B34"] = "Trades not Ready"
        wb_assets["B35"] = "Trades Unavailable"
        wb_assets["B36"] = "VF Checklist Issuex"

        # passdown_assets --> KAWI and SMA
        wb_assets["C1"] = "Yes"
        wb_assets["C2"] = "No"
        wb_assets["C3"] = "N/A"

        #Set up data validation for for required cells in tables
        #SMA and KAWI data validation
        dv1 = DataValidation(type="list", formula1="passdown_assets!$C$1:$C$3", allow_blank=True)
        dv1.add("E1:F103")

        #Set up data validation for AIO Request
        dv2 = DataValidation(type="list", formula1="passdown_assets!$A$1:$A$26", allow_blank=True)
        dv2.add("H1:H103")

        #Set up data validation for AIO Completion
        dv3 = DataValidation(type="list", formula1="passdown_assets!$B$1:$B$36", allow_blank=True)
        dv3.add("I1:I103")

        #add data validation to the sheet
        sheet.add_data_validation(dv1)
        sheet.add_data_validation(dv2)
        sheet.add_data_validation(dv3)

        #Text Wrapping in tables
        for row in sheet.iter_rows(min_row=5, max_row=103, min_col=1, max_col=16):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')

        #Add look ahead key
        sheet["P4"].fill = PatternFill(patternType="solid", fgColor="00808080")
        sheet["P5"].fill = PatternFill(patternType="solid", fgColor="0000FF00") # Green - Sheduled
        sheet["P5"] = "Scheduled"
        sheet["P6"].fill = PatternFill(patternType="solid", fgColor="00FF00FF") # Pink - Awaiting next steps
        sheet["P6"] = "Awaiting next steps"
        sheet["P7"].fill = PatternFill(patternType="solid", fgColor="00FF0000") # Red - Immediate Escalation (High Priority) 
        sheet["P7"] = "Immediate Escalation (High Priority)"
        sheet["P8"].fill = PatternFill(patternType="solid", fgColor="00FFFF00")  # Yellow - Waiting on Parts
        sheet["P8"] = "Waiting on Parts"
        for row in sheet.iter_rows(min_row=9, max_row=103, min_col=16, max_col=16):
            for cell in row:
                cell.fill = PatternFill(patternType="solid", fgColor="00808080")

        #Add contents sheet link
        sheet["P10"] = "Contents"
        sheet["P10"].hyperlink = "#contents!A1"
        sheet["P10"].font = Font(
            bold=True, color="000000FF", underline="single", size=12)

        #Add passdown summary sheet link
        sheet["P11"] = "Passdown Summary"
        sheet["P11"].hyperlink = "#passdown_summary!A1"
        sheet["P11"].font = Font(
            bold=True, color="000000FF", underline="single", size=12)

        #Add passdown close out 'button'
        sheet["P12"] = "Close Out"
        sheet["P12"].font = Font(
            bold=True, color="000000FF", underline="single", size=12)
        #TODO add close out macro to button

        #Fix bottom border on last row
        sheet["P1"].border = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(border_style="thick"))
        sheet["P2"].border = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(border_style="thick"))
        sheet["P3"].border = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(border_style="thick"))

    # Save the modified workbook
    wb.save(filename)

def create_contents_sheet(filename):

    '''
    Create a contents sheet in the workbook with links to all other sheets

    :param filename: a string containing the name of the file to load the workbook from (including the file extension)
    :requires: openpyxl
    :date: 2023-03-22
    '''
    # Load the workbook
    wb = openpyxl.load_workbook(filename)

    # Create a new sheet for the contents
    contents_sheet = wb.worksheets[0]
    contents_sheet.column_dimensions["A"].width = 20

    # Add the title
    contents_sheet["A1"] = "Links to sheets"
    contents_sheet["A1"].font = Font(bold=True, size=16, color="00C0C0C0")
    contents_sheet["A1"].alignment = Alignment(horizontal="center")
    contents_sheet["A1"].fill = PatternFill(patternType="solid", fgColor="000000")

    # Loop over the sheets and add links to the contents sheet
    #TODO add weekday, workweek and headers/formatting
    for sheet in wb.sheetnames:
        # Skip the contents sheet and workbook assets
        if sheet == "Contents":
            continue
        if sheet == "passdown_assets":
            continue

        # Add a hyperlink to the sheet
        cell = contents_sheet.cell(row=contents_sheet.max_row+1, column=1)
        cell.value = sheet
        cell.hyperlink = f"#'{sheet}'!A1"
        cell.font = Font(color="0000FF", underline="single", size=12, bold=True)

        if cell.row % 2 == 0:
            cell.fill = PatternFill(patternType="solid", fgColor="C0C0C0")

    # Save the modified workbook
    wb.save(filename)

project_workbook = "excel/Project_Passdown_WW14-WW28.xlsx"
my_workdays = get_workdays(start_day, end_day)
sheet_names = get_sheet_names(my_workdays)
create_workbook(sheet_names, project_workbook)
create_contents_sheet(project_workbook)
create_daily_sheets(project_workbook, my_workdays)

